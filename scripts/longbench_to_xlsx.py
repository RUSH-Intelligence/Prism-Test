#!/usr/bin/env python
"""Fill a per-model LongBench sheet in ``Ridge Press.xlsx`` from a sweep manifest.

Reads the ``manifest.json`` written by ``scripts/longbench_sweep.py`` and each
referenced ``metrics.json`` (produced by ``LongBenchBenchmark.score``), then
writes a new sheet laid out like the RULER sheets but without the context-length
axis (LongBench has none):

    LongBench  <model>
    Full         <16 task scores>                      Avg
    0.2  | NrtvQA Qasper MF-en ... RB-P | Avg
      Knorm        ...
      CurDkv       ...
      ... (9 methods)
    0.4  | ...
    ...

Existing sheets (the RULER tabs) are left untouched.

Usage
-----
    python scripts/longbench_to_xlsx.py
    python scripts/longbench_to_xlsx.py --manifest results/longbench_sweep/<model>/manifest.json
    python scripts/longbench_to_xlsx.py --xlsx "Ridge Press.xlsx" --sheet LongBench-Llama-3.1-8B
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parent.parent

# Subset name -> paper Table 1 column header.
SUBSET_TO_HEADER = {
    "narrativeqa": "NrtvQA", "qasper": "Qasper", "multifieldqa_en": "MF-en",
    "hotpotqa": "HotpotQA", "2wikimqa": "2WikiMQA", "musique": "Musique",
    "gov_report": "GovReport", "qmsum": "QMSum", "multi_news": "MultiNews",
    "trec": "TREC", "triviaqa": "TriviaQA", "samsum": "SAMSum",
    "passage_count": "PCount", "passage_retrieval_en": "PRe",
    "lcc": "LCC", "repobench-p": "RB-P",
}

# Row order for each ratio block (the superset; Full is rendered once on top).
# Ridge is swept over envelope_gamma ∈ {0.0, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0} —
# each gamma gets its own row. H2O is intentionally excluded (see
# scripts/longbench_sweep.py for why).
METHOD_ORDER = [
    "Knorm", "CurDkv", "PyramidKv", "SnapKv", "Expected_attn",
    "Key_Diff", "Compactor", "StreamingLLM",
    "Ridge_g0.0", "Ridge_g0.5", "Ridge_g1.0", "Ridge_g1.5",
    "Ridge_g2.0", "Ridge_g2.5", "Ridge_g3.0",
]


def _row_label(cell: dict) -> str:
    """The display label for the row — cell_id stripped of the ``__r<ratio>`` suffix.

    Falls back to ``label`` for the Full baseline (which has no ratio suffix)."""
    cid = cell.get("cell_id") or cell["label"]
    return cid.split("__r", 1)[0]


def load_results(manifest: dict) -> dict:
    """Return {(row_label, ratio_or_None): {subset: score}} from the manifest."""
    out: dict = {}
    for cell in manifest["cells"]:
        if not cell.get("metrics"):
            continue
        try:
            data = json.loads(Path(cell["metrics"]).read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        out[(_row_label(cell), cell["ratio"])] = data.get("task_scores", {})
    return out


def _avg(values: list) -> float | None:
    nums = [v for v in values if isinstance(v, (int, float))]
    return round(sum(nums) / len(nums), 2) if nums else None


def write_sheet(wb, sheet_name: str, model: str, subsets: list[str],
                ratios: list[float], results: dict) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)
    headers = [SUBSET_TO_HEADER.get(s, s) for s in subsets]

    def write_method_row(r: int, label: str, ratio: float | None) -> None:
        scores = results.get((label, ratio), {})
        ws.cell(r, 1, label)
        vals = []
        for c, sub in enumerate(subsets, start=2):
            v = scores.get(sub)
            vals.append(v)
            if v is not None:
                ws.cell(r, c, round(float(v), 2))
        a = _avg(vals)
        if a is not None:
            ws.cell(r, len(subsets) + 2, a)

    row = 1
    ws.cell(row, 1, "LongBench").font = bold
    ws.cell(row, 2, model).font = bold
    row += 1

    # Full baseline (once).
    write_method_row(row, "Full", None)
    ws.cell(row, 1).font = bold
    row += 2

    # One block per compression ratio.
    for ratio in ratios:
        ws.cell(row, 1, ratio).font = bold
        for c, h in enumerate(headers, start=2):
            ws.cell(row, c, h).font = bold
        ws.cell(row, len(subsets) + 2, "Avg").font = bold
        row += 1
        for label in METHOD_ORDER:
            write_method_row(row, label, ratio)
            row += 1
        row += 1  # blank separator


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--manifest", required=False,
                    help="Path to sweep manifest.json (default: newest under results/longbench_sweep)")
    ap.add_argument("--xlsx", default=str(REPO_ROOT / "Ridge Press.xlsx"))
    ap.add_argument("--sheet", default=None, help="Sheet name (default: LongBench-<model-slug>)")
    args = ap.parse_args()

    if args.manifest:
        manifest_path = Path(args.manifest)
    else:
        hits = sorted((REPO_ROOT / "results" / "longbench_sweep").rglob("manifest.json"),
                      key=lambda p: p.stat().st_mtime)
        if not hits:
            raise SystemExit("No manifest.json found under results/longbench_sweep — run the sweep first.")
        manifest_path = hits[-1]

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    model = manifest["model"]
    subsets = manifest["subsets"]
    ratios = manifest["ratios"]
    results = load_results(manifest)

    sheet = args.sheet or f"LongBench-{model.split('/')[-1]}"
    xlsx = Path(args.xlsx)
    if xlsx.exists():
        wb = openpyxl.load_workbook(xlsx)
        action = "updated"
    else:
        # Fresh workbook — drop the empty default sheet so the LongBench tab is first.
        wb = openpyxl.Workbook()
        default = wb.active
        if default is not None and default.max_row == 1 and default.max_column == 1 \
                and default.cell(1, 1).value is None:
            wb.remove(default)
        action = "created"
    write_sheet(wb, sheet, model, subsets, ratios, results)
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx)

    filled = len(results)
    total = 1 + len(METHOD_ORDER) * len(ratios)
    print(f"Wrote sheet '{sheet}' to {xlsx} ({action}; {filled}/{total} cells with data).")
    missing = [f"{lbl}@{ratio}" for ratio in ratios for lbl in METHOD_ORDER
               if (lbl, ratio) not in results]
    if (("Full", None) not in results):
        missing.insert(0, "Full")
    if missing:
        print(f"Missing cells (left blank): {', '.join(missing)}")


if __name__ == "__main__":
    main()
